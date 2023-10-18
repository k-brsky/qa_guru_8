from zipfile import ZipFile, ZIP_DEFLATED
import os
from os.path import basename
from openpyxl import load_workbook
from xlrd import open_workbook

def test_make_zip():
    path = 'resources'
    file_dir = os.listdir(path)
    os.mkdir('tmp')
    with ZipFile('tmp/homework.zip', mode='w', compression=ZIP_DEFLATED) as zf:
        for file in file_dir:
            add_file = os.path.join(path, file)
            zf.write(add_file, basename(add_file))

def test_check_txt():
    with open('resources/file_txt.txt') as f:
        text_resources = f.read()

    with ZipFile('tmp/homework.zip') as zf:
        text_zip = zf.read('file_txt.txt').decode()
        assert text_resources == text_zip

def test_check_xlsx():
    with open('resources/test_sheet_xlsx.xlsx') as f:
        wb_resources = load_workbook('resources/test_sheet_xlsx.xlsx')
        sheet_resources = wb_resources.active
        cell_name_resources = sheet_resources.cell(row = 3, column = 1).value
        cell_age_resources = sheet_resources.cell(row = 3, column = 2).value

    with ZipFile('tmp/homework.zip') as zf:
        wb_zip = load_workbook(zf.open('test_sheet_xlsx.xlsx'))
        sheet_zip = wb_zip.active
        cell_name_zip = sheet_zip.cell(row = 3, column = 1).value
        cell_age_zip = sheet_zip.cell(row = 3, column = 2).value
        assert cell_name_resources == cell_name_zip
        assert cell_age_resources == cell_age_zip

def test_check_xls():
    work_book_resources = open_workbook('resources/test_sheet_xls.xls')
    cell_animal_resorces = work_book_resources.sheet_by_index(0).cell_value(3, 0)
    cell_count_legs_resources = work_book_resources.sheet_by_index(0).cell_value(3, 1)

    with ZipFile('tmp/homework.zip') as zf:
        work_book_zip = open_workbook(zf.open('test_sheet_xls.xls'))
        cell_animal_zip = work_book_zip.sheet_by_index(0).cell_value(3, 0)
        cell_count_legs_zip = work_book_zip.sheet_by_index(0).cell_value(3, 1)
        assert cell_animal_resorces == cell_animal_zip
        assert cell_count_legs_resources == cell_count_legs_zip









